import pandas as pd
import configparser
import ast
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dfcx_scrapi.core.conversation import DialogflowConversation


# -------------------------------------------------
# Helpers
# -------------------------------------------------
def parse_params_maybe_dict(value):

    if value is None:
        return None

    if isinstance(value, dict):
        return value

    if isinstance(value, str):
        s = value.strip()

        if s == "":
            return None

        try:
            d = json.loads(s)
            if isinstance(d, dict):
                return d
        except Exception:
            pass

        try:
            d = ast.literal_eval(s)
            if isinstance(d, dict):
                return d
        except Exception:
            return None

    return None


def first_nonempty_value(d):

    if not isinstance(d, dict):
        return None

    for v in d.values():

        if v is None:
            continue

        if str(v).strip() == "":
            continue

        return v

    return None


def normalize_text(x):

    if pd.isna(x):
        return ""

    return str(x).strip().lower().replace(" ", "")


def is_empty_params_cell(v):

    if v is None:
        return True

    try:
        if pd.isna(v):
            return True
    except Exception:
        pass

    if isinstance(v, str):

        s = v.strip()

        if s == "":
            return True

        d = parse_params_maybe_dict(s)

        if isinstance(d, dict) and len(d) == 0:
            return True

    if isinstance(v, dict):
        return len(v) == 0

    return False


# -------------------------------------------------
# Load Config
# -------------------------------------------------
config = configparser.ConfigParser()
config.read("config.properties")

agent_id = config.get("dialogflow", "agent_id")
creds_path = config.get("dialogflow", "creds_path")

excel_file = config.get("input", "excel_file")

start_sheet = config.getint("input", "start_sheet")
end_sheet = config.getint("input", "end_sheet")

flow_display_name = config.get(
    "input",
    "flow_display_name"
)

page_display_name = config.get(
    "input",
    "page_display_name"
)

Utterances = config.get(
    "input",
    "Utterances"
)

intent_column = config.get(
    "input",
    "intent_column"
)

# Existing columns in the sheet that will be FILLED (not added)
detected_column = config.get(
    "input",
    "detected_column",
    fallback="Actual Intent"
)

result_column = config.get(
    "input",
    "result_column",
    fallback="Result"
)

# Results are written back into the SAME file
result_file = excel_file

# -------------------------------------------------
# Initialize Conversation
# -------------------------------------------------
conversation = DialogflowConversation(
    agent_id=agent_id
)

# -------------------------------------------------
# Open workbook (keeps ALL original columns,
# sheets, and formatting intact)
# -------------------------------------------------
wb = load_workbook(excel_file)

summary_sheet_name = "Summary"

# Sheet numbering EXCLUDES the Summary sheet, so
# start_sheet / end_sheet always refer to the same
# data sheets on every run (1 = first data sheet)
sheet_names = [
    s for s in wb.sheetnames
    if s != summary_sheet_name
]

total_sheets = len(sheet_names)

if start_sheet < 1:
    raise ValueError(
        "start_sheet must be >= 1"
    )

if end_sheet > total_sheets:
    raise ValueError(
        f"Workbook contains only "
        f"{total_sheets} data sheets."
    )

# -------------------------------------------------
# Process Sheet by Sheet
# -------------------------------------------------
for sheet_num in range(
    start_sheet,
    end_sheet + 1
):

    sheet_name = sheet_names[sheet_num - 1]

    print(
        f"\nProcessing Sheet "
        f"{sheet_num}: {sheet_name}"
    )

    df = pd.read_excel(
        excel_file,
        sheet_name=sheet_name
    )

    for col in [
        Utterances,
        intent_column,
        detected_column,
        result_column
    ]:

        if col not in df.columns:
            raise ValueError(
                f"Column '{col}' not found "
                f"in sheet '{sheet_name}'"
            )

    df[Utterances] = (
        df[Utterances]
        .astype(str)
        .replace({"nan": ""})
        .fillna("")
        .str.strip()
    )

    # Keep only rows with an actual utterance.
    # This skips blank rows and the "Accuracy"
    # row added at the bottom by a previous run.
    valid_mask = df[Utterances] != ""

    df_valid = df[valid_mask]

    # Excel row for each valid data row
    # (header is row 1, data starts at row 2)
    valid_excel_rows = [
        idx + 2 for idx in df_valid.index
    ]

    if len(df_valid) == 0:
        print(
            f"No utterances found in "
            f"'{sheet_name}', skipping."
        )
        continue

    test_set = pd.DataFrame(
        {
            "flow_display_name":
                flow_display_name,

            "page_display_name":
                page_display_name,

            "utterance":
                df_valid[Utterances].values,

            "inject_parameters":
                "",

            "end_user_metadata":
                ""
        }
    )

    results = conversation.run_intent_detection(
        test_set,
        100,
        10
    )

    if "detected_intent" not in results.columns:
        raise ValueError(
            "detected_intent column missing."
        )

    if "confidence" not in results.columns:
        raise ValueError(
            "confidence column missing."
        )

    expected = df_valid[intent_column].reset_index(
        drop=True
    )

    # -------------------------------------
    # Intent Match
    # -------------------------------------
    string_match = (
        expected
        .astype(str)
        .apply(normalize_text)
        .values
        ==
        results["detected_intent"]
        .astype(str)
        .apply(normalize_text)
        .values
    )

    conf_ok = (
        results["confidence"]
        .astype(float)
        .values
        >= 0.30
    )

    intent_match = string_match & conf_ok

    # -------------------------------------
    # Entity Extraction
    # -------------------------------------
    params_col = (
        "parameters_set"
        if "parameters_set" in results.columns
        else (
            "parameters"
            if "parameters" in results.columns
            else None
        )
    )

    if params_col is None:

        results["parameters_set"] = None
        params_col = "parameters_set"

    extracted_entity = (
        results[params_col]
        .apply(
            lambda v: first_nonempty_value(
                parse_params_maybe_dict(v)
            )
        )
    )

    entity_match = (
        extracted_entity
        .apply(normalize_text)
        .values
        ==
        expected
        .apply(normalize_text)
        .values
    ) & conf_ok

    first5_empty = (
        results[params_col]
        .head(5)
        .apply(is_empty_params_cell)
        .all()
    )

    # -------------------------------------
    # Decide what goes into the existing
    # "Actual Intent" and "Result" columns
    # -------------------------------------
    if first5_empty:
        # Intent mode
        actual_values = (
            results["detected_intent"]
            .astype(str)
            .tolist()
        )
        match_flags = intent_match
    else:
        # Entity mode
        actual_values = [
            "" if v is None else str(v)
            for v in extracted_entity
        ]
        match_flags = entity_match

    result_values = [
        "Pass" if m else "Fail"
        for m in match_flags
    ]

    total_count = len(match_flags)
    pass_count = int(match_flags.sum())

    pass_pct = (
        round((pass_count / total_count) * 100, 2)
        if total_count > 0
        else 0
    )

    print(f"Accuracy: {pass_pct:.2f}%")

    # -------------------------------------
    # Write values into the EXISTING columns
    # of the original sheet (no new columns)
    # -------------------------------------
    ws = wb[sheet_name]

    header = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    detected_col_idx = header[detected_column]
    result_col_idx = header[result_column]

    for i in range(len(actual_values)):

        excel_row = valid_excel_rows[i]

        ws.cell(
            row=excel_row,
            column=detected_col_idx,
            value=actual_values[i]
        )

        ws.cell(
            row=excel_row,
            column=result_col_idx,
            value=result_values[i]
        )

    # -------------------------------------
    # Accuracy row at the bottom of the
    # Result column
    # -------------------------------------
    accuracy_row = max(valid_excel_rows) + 1

    label_cell = ws.cell(
        row=accuracy_row,
        column=detected_col_idx,
        value="Accuracy"
    )
    label_cell.font = Font(bold=True)

    acc_cell = ws.cell(
        row=accuracy_row,
        column=result_col_idx,
        value=f"{pass_pct:.2f}%"
    )
    acc_cell.font = Font(bold=True)

# -------------------------------------------------
# Recalculate Summary from ALL data sheets.
# Because results are saved in the same file,
# this picks up sheets tested in PREVIOUS runs
# too (e.g. run 1-15, then run 16-20).
# -------------------------------------------------
summary_rows = []

for sheet_name in sheet_names:

    ws = wb[sheet_name]

    header = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    if result_column not in header:
        continue

    result_col_idx = header[result_column]

    pass_count = 0
    fail_count = 0

    for row in ws.iter_rows(
        min_row=2,
        min_col=result_col_idx,
        max_col=result_col_idx
    ):

        val = row[0].value

        if val is None:
            continue

        val = str(val).strip().lower()

        if val == "pass":
            pass_count += 1
        elif val == "fail":
            fail_count += 1
        # anything else (e.g. the accuracy
        # "% value" cell) is ignored

    total_count = pass_count + fail_count

    # Only include sheets that have results
    if total_count == 0:
        continue

    pass_pct = round(
        (pass_count / total_count) * 100, 2
    )
    fail_pct = round(
        (fail_count / total_count) * 100, 2
    )

    summary_rows.append(
        [
            sheet_name,
            total_count,
            pass_count,
            fail_count,
            pass_pct,
            fail_pct
        ]
    )

# -------------------------------------------------
# Create Summary sheet as the FIRST sub sheet
# -------------------------------------------------
# Remove old Summary sheet if it already exists
if summary_sheet_name in wb.sheetnames:
    del wb[summary_sheet_name]

ws_summary = wb.create_sheet(
    summary_sheet_name,
    0  # index 0 = first sheet
)

summary_headers = [
    "Sheet Name",
    "Total Count",
    "Pass Count",
    "Fail Count",
    "Pass Percentage",
    "Fail Percentage"
]

header_font = Font(bold=True)

center_align = Alignment(
    horizontal="center",
    vertical="center"
)

thin = Side(style="thin")

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

for col_idx, header_text in enumerate(
    summary_headers,
    start=1
):

    cell = ws_summary.cell(
        row=1,
        column=col_idx,
        value=header_text
    )

    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

for row_idx, row_data in enumerate(
    summary_rows,
    start=2
):

    for col_idx, value in enumerate(
        row_data,
        start=1
    ):

        cell = ws_summary.cell(
            row=row_idx,
            column=col_idx,
            value=value
        )

        # Numbers right-aligned, sheet name left
        if col_idx == 1:
            cell.alignment = Alignment(
                horizontal="left"
            )
        else:
            cell.alignment = Alignment(
                horizontal="right"
            )

# Auto-fit-ish column widths
for col_idx, header_text in enumerate(
    summary_headers,
    start=1
):

    max_len = len(header_text)

    for row_data in summary_rows:

        val_len = len(str(row_data[col_idx - 1]))

        if val_len > max_len:
            max_len = val_len

    ws_summary.column_dimensions[
        get_column_letter(col_idx)
    ].width = max_len + 4

# Make Summary the active sheet on open
wb.active = 0

# -------------------------------------------------
# Save workbook with filled columns + Summary
# -------------------------------------------------
wb.save(result_file)

print("\nCompleted Successfully")
print(f"Output File    : {result_file}")
